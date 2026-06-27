import uuid
import json
import io
from datetime import datetime, timedelta
from typing import Optional
import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel

from ..db.database import get_connection
from ..science.ict_engine import compute_ict

router = APIRouter(tags=["excel"])

# In-memory session store: token -> {bytes, created_at}
EXCEL_SESSIONS: dict[str, dict] = {}
SESSION_TTL_MINUTES = 30


def _get_session(token: str) -> dict:
    session = EXCEL_SESSIONS.get(token)
    if not session:
        raise HTTPException(404, "Session not found or expired. Please re-upload the file.")
    if datetime.utcnow() - session["created_at"] > timedelta(minutes=SESSION_TTL_MINUTES):
        del EXCEL_SESSIONS[token]
        raise HTTPException(410, "Session expired. Please re-upload the file.")
    return session


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


# ── Upload ────────────────────────────────────────────────────────────────────

@router.post("/api/excel/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported.")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Failed to open Excel file: {e}")

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "row_count": ws.max_row or 0})
    wb.close()

    token = str(uuid.uuid4())
    EXCEL_SESSIONS[token] = {"bytes": content, "created_at": datetime.utcnow()}
    return {"token": token, "sheets": sheets}


# ── Raw sheet preview ─────────────────────────────────────────────────────────

class SheetRawRequest(BaseModel):
    sheet_name: str


ROW_CAP = 500


@router.post("/api/excel/{token}/sheet-raw")
def sheet_raw(token: str, body: SheetRawRequest):
    session = _get_session(token)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(session["bytes"]), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Failed to read workbook: {e}")

    if body.sheet_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(404, f"Sheet '{body.sheet_name}' not found.")

    ws = wb[body.sheet_name]
    total_rows = ws.max_row or 0
    rows = []
    for i, row in enumerate(ws.iter_rows()):
        if i >= ROW_CAP:
            break
        rows.append([_cell_str(cell.value) for cell in row])
    wb.close()

    result = {"rows": rows, "row_count": total_rows}
    if total_rows > ROW_CAP:
        result["truncated"] = True
        result["total_rows"] = total_rows
    return result


# ── Import ────────────────────────────────────────────────────────────────────

class SampleImportConfig(BaseModel):
    sheet_name: str
    sample_name: str
    header_row_index: int
    data_row_indices: list[int]
    column_map: dict[str, str]
    group_column: Optional[str] = None


class ImportRequest(BaseModel):
    project_id: str
    samples: list[SampleImportConfig]


def _insert_sample(conn, project_id: str, name: str, column_map: dict, rows: list[list[str]], headers: list[str]) -> dict:
    """Insert one sample + observations. Returns {sample_id, name, obs_count}."""
    col_index = {h: i for i, h in enumerate(headers)}

    time_col = column_map.get("time")
    conc_col = column_map.get("concentration")
    cfu_col = column_map.get("cfu")
    ict_col = column_map.get("ict")
    repl_col = column_map.get("replicate")
    dose_col = column_map.get("dose")

    if not time_col or not conc_col or not cfu_col:
        raise ValueError("Column map must include time, concentration, and cfu.")

    def get_col(row, col_name):
        if col_name and col_name in col_index:
            v = row[col_index[col_name]]
            return v if v != "" else None
        return None

    times, concs, cfus, icts_raw, repls, doses = [], [], [], [], [], []
    for row in rows:
        t = get_col(row, time_col)
        c = get_col(row, conc_col)
        f = get_col(row, cfu_col)
        if t is None or c is None or f is None:
            continue
        try:
            times.append(float(t))
            concs.append(float(c))
            cfus.append(float(f))
            icts_raw.append(float(get_col(row, ict_col)) if ict_col and get_col(row, ict_col) is not None else None)
            repls.append(get_col(row, repl_col))
            doses.append(float(get_col(row, dose_col)) if dose_col and get_col(row, dose_col) is not None else None)
        except (ValueError, TypeError):
            continue

    if not times:
        raise ValueError("No valid numeric rows found with the given column mapping.")

    icts = []
    for i, iv in enumerate(icts_raw):
        icts.append(iv if iv is not None else None)
    computed = compute_ict(times, concs)
    icts = [iv if iv is not None else computed[i] for i, iv in enumerate(icts)]

    sample_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO samples (id, project_id, name, column_map) VALUES (?, ?, ?, ?)",
        (sample_id, project_id, name, json.dumps(column_map)),
    )
    for i, (t, c, f, ict, repl, dose) in enumerate(zip(times, concs, cfus, icts, repls, doses)):
        conn.execute(
            "INSERT INTO observations (sample_id, time, concentration, cfu, ict, replicate, dose, row_index) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (sample_id, t, c, f, ict, repl, dose, i),
        )
    obs_count = len(times)
    return {"sample_id": sample_id, "name": name, "obs_count": obs_count}


@router.post("/api/excel/{token}/import")
def import_excel(token: str, body: ImportRequest):
    session = _get_session(token)
    conn = get_connection()

    proj = conn.execute("SELECT id FROM projects WHERE id = ?", (body.project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, "Project not found.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(session["bytes"]), read_only=True, data_only=True)
    except Exception as e:
        conn.close()
        raise HTTPException(400, f"Failed to read workbook: {e}")

    # Pre-read all needed sheets into a dict of row lists
    sheet_rows: dict[str, list[list[str]]] = {}
    for cfg in body.samples:
        if cfg.sheet_name not in sheet_rows:
            if cfg.sheet_name not in wb.sheetnames:
                sheet_rows[cfg.sheet_name] = []
                continue
            ws = wb[cfg.sheet_name]
            sheet_rows[cfg.sheet_name] = [[_cell_str(cell.value) for cell in row] for row in ws.iter_rows()]
    wb.close()

    imported = []
    errors = []

    for cfg in body.samples:
        all_rows = sheet_rows.get(cfg.sheet_name, [])
        if not all_rows:
            errors.append({"sample_name": cfg.sample_name, "reason": f"Sheet '{cfg.sheet_name}' not found or empty."})
            continue

        # Validate indices
        max_idx = len(all_rows) - 1
        if cfg.header_row_index < 0 or cfg.header_row_index > max_idx:
            errors.append({"sample_name": cfg.sample_name, "reason": f"Header row index {cfg.header_row_index} out of range."})
            continue

        headers = all_rows[cfg.header_row_index]
        valid_data_indices = [i for i in cfg.data_row_indices if 0 <= i <= max_idx and i != cfg.header_row_index]
        data_rows = [all_rows[i] for i in valid_data_indices]

        if not data_rows:
            errors.append({"sample_name": cfg.sample_name, "reason": "No data rows selected."})
            continue

        if cfg.group_column:
            # Split by group column
            group_col_name = cfg.group_column
            if group_col_name not in headers:
                errors.append({"sample_name": cfg.sample_name, "reason": f"Group column '{group_col_name}' not found in headers."})
                continue

            col_index = {h: i for i, h in enumerate(headers)}
            gi = col_index[group_col_name]
            groups: dict[str, list] = {}
            null_count = 0
            for row in data_rows:
                gval = row[gi] if gi < len(row) else ""
                if gval == "":
                    null_count += 1
                    continue
                groups.setdefault(gval, []).append(row)

            if null_count > 0:
                errors.append({"sample_name": cfg.sample_name, "reason": f"Warning: {null_count} rows skipped (blank group column value)."})

            for gval, grows in groups.items():
                gname = f"{cfg.sample_name} – {gval}"
                try:
                    result = _insert_sample(conn, body.project_id, gname, cfg.column_map, grows, headers)
                    conn.commit()
                    imported.append(result)
                except Exception as e:
                    conn.rollback()
                    errors.append({"sample_name": gname, "reason": str(e)})
        else:
            try:
                result = _insert_sample(conn, body.project_id, cfg.sample_name, cfg.column_map, data_rows, headers)
                conn.commit()
                imported.append(result)
            except Exception as e:
                conn.rollback()
                errors.append({"sample_name": cfg.sample_name, "reason": str(e)})

    conn.close()
    return {"imported": imported, "errors": errors}


# ── Session cleanup ───────────────────────────────────────────────────────────

@router.delete("/api/excel/{token}")
def delete_session(token: str):
    EXCEL_SESSIONS.pop(token, None)
    return {"ok": True}


# ── Import templates ──────────────────────────────────────────────────────────

class TemplateCreate(BaseModel):
    name: str
    column_map: dict[str, str]
    group_column: Optional[str] = None


@router.get("/api/templates")
def list_templates():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM import_templates ORDER BY created_at DESC").fetchall()
    conn.close()
    result = []
    for r in rows:
        t = dict(r)
        t["column_map"] = json.loads(t["column_map"]) if t["column_map"] else {}
        result.append(t)
    return result


@router.post("/api/templates")
def create_template(body: TemplateCreate):
    conn = get_connection()
    tid = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO import_templates (id, name, column_map, group_column) VALUES (?, ?, ?, ?)",
        (tid, body.name, json.dumps(body.column_map), body.group_column),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM import_templates WHERE id = ?", (tid,)).fetchone()
    conn.close()
    t = dict(row)
    t["column_map"] = json.loads(t["column_map"])
    return t


@router.delete("/api/templates/{template_id}")
def delete_template(template_id: str):
    conn = get_connection()
    conn.execute("DELETE FROM import_templates WHERE id = ?", (template_id,))
    conn.commit()
    conn.close()
    return {"ok": True}
