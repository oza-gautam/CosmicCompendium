import io
import json
import uuid
from typing import Optional
import openpyxl
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel

from ..db.database import get_connection
from ..science.ict_engine import compute_ict

router = APIRouter(tags=["quick_import"])

# ── Column name heuristics ────────────────────────────────────────────────────

_TIME_NAMES = {"time", "contact_time", "minute", "minutes", "t", "time_min"}
_CFU_NAMES = {"cfu", "cfu/ml", "enterococcus", "count", "colony", "log_cfu", "organisms", "n"}
_CONC_NAMES = {"concentration", "conc", "residual", "paa", "paa_residual", "cl2", "chlorine", "disinfectant", "mg/l"}
_GROUP_NAMES = {"sample", "dose", "group", "id", "treatment", "paa dose (mg/l)", "sample_id"}


def _normalize(s: str) -> str:
    return s.lower().strip().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace("-", "_")


def _detect_col_role(header: str) -> Optional[str]:
    n = _normalize(header)
    if n in _TIME_NAMES or any(t in n for t in ("time", "minute", "contact")):
        return "time"
    if n in _CFU_NAMES or any(t in n for t in ("cfu", "enterococcus", "colony")):
        return "cfu"
    if n in _CONC_NAMES or any(t in n for t in ("residual", "paa", "chlorine", "conc", "mg_l")):
        return "concentration"
    if n in _GROUP_NAMES or any(t in n for t in ("sample", "dose", "group")):
        return "group"
    return None


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def _is_numeric(val: str) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False


# ── Pydantic models ───────────────────────────────────────────────────────────

class DetectedColMap(BaseModel):
    group: Optional[str] = None
    time: Optional[str] = None
    cfu: Optional[str] = None
    concentration: Optional[str] = None


class SampleGroupPreview(BaseModel):
    group_value: str
    sample_name: str
    row_count: int


class SheetPreview(BaseModel):
    sheet_name: str
    experiment_name: str
    header_row_index: int
    col_map: DetectedColMap
    sample_groups: list[SampleGroupPreview]
    detected: bool
    raw_rows: Optional[list[list[str]]] = None  # only when detected=False
    all_headers: list[str]


class QuickImportPreview(BaseModel):
    sheets: list[SheetPreview]


class ConfirmedColMap(BaseModel):
    group: Optional[str] = None
    time: str
    cfu: str
    concentration: str


class SheetMetadata(BaseModel):
    experiment_date: Optional[str] = None  # YYYY-MM-DD
    analyst: Optional[str] = None
    notes: Optional[str] = None


class SheetImportConfig(BaseModel):
    sheet_name: str
    experiment_name: str
    header_row_index: int
    col_map: ConfirmedColMap
    data_row_indices: Optional[list[int]] = None  # None = auto (all non-header rows)
    metadata: Optional[SheetMetadata] = None


class QuickImportExecuteRequest(BaseModel):
    project_id: str
    sheets: list[SheetImportConfig]


class QuickImportResult(BaseModel):
    experiment_ids: list[int]
    first_experiment_id: int
    total_samples: int
    total_observations: int


# ── Detection logic ───────────────────────────────────────────────────────────

def _detect_sheet(ws, sheet_name: str) -> SheetPreview:
    all_sheet_rows = [[_cell_str(cell.value) for cell in row] for row in ws.iter_rows()]

    # Find header row: first row where majority of cells are non-numeric strings
    header_row_index = 0
    for i, row in enumerate(all_sheet_rows[:5]):
        non_empty = [c for c in row if c]
        if not non_empty:
            continue
        non_numeric = [c for c in non_empty if not _is_numeric(c)]
        if len(non_numeric) >= max(1, len(non_empty) // 2):
            header_row_index = i
            break

    headers = all_sheet_rows[header_row_index] if all_sheet_rows else []

    col_map = DetectedColMap()
    for h in headers:
        if not h:
            continue
        role = _detect_col_role(h)
        if role == "group" and col_map.group is None:
            col_map.group = h
        elif role == "time" and col_map.time is None:
            col_map.time = h
        elif role == "cfu" and col_map.cfu is None:
            col_map.cfu = h
        elif role == "concentration" and col_map.concentration is None:
            col_map.concentration = h

    detected = bool(col_map.time and col_map.cfu and col_map.concentration)

    sample_groups: list[SampleGroupPreview] = []
    if detected and col_map.group:
        group_idx = next((i for i, h in enumerate(headers) if h == col_map.group), None)
        if group_idx is not None:
            seen: dict[str, int] = {}
            for row in all_sheet_rows[header_row_index + 1:]:
                if group_idx >= len(row):
                    continue
                gval = row[group_idx].strip()
                if not gval or not _is_numeric(gval):
                    continue
                seen[gval] = seen.get(gval, 0) + 1
            for gval, count in seen.items():
                sample_groups.append(SampleGroupPreview(
                    group_value=gval,
                    sample_name=f"{sheet_name} – {gval} mg/L",
                    row_count=count,
                ))

    raw_rows = None if detected else all_sheet_rows

    return SheetPreview(
        sheet_name=sheet_name,
        experiment_name=sheet_name,
        header_row_index=header_row_index,
        col_map=col_map,
        sample_groups=sample_groups,
        detected=detected,
        raw_rows=raw_rows,
        all_headers=headers,
    )


# ── Insertion helper ──────────────────────────────────────────────────────────

def _insert_sample_quick(conn, project_id: str, exp_id: int, sample_name: str,
                          col_map: ConfirmedColMap, data_rows: list[list[str]],
                          headers: list[str]) -> int:
    col_index = {h: i for i, h in enumerate(headers)}

    def ci(col_name: Optional[str]) -> Optional[int]:
        if col_name and col_name in col_index:
            return col_index[col_name]
        return None

    time_i = ci(col_map.time)
    cfu_i = ci(col_map.cfu)
    conc_i = ci(col_map.concentration)

    if time_i is None or cfu_i is None or conc_i is None:
        raise ValueError(f"Could not find required columns in headers {headers}")

    times, concs, cfus = [], [], []
    bad_rows = []
    for row_num, row in enumerate(data_rows):
        def get(idx):
            return row[idx] if idx is not None and idx < len(row) else ""
        t, c, f = get(time_i), get(conc_i), get(cfu_i)
        if not t or not c or not f:
            continue
        try:
            times.append(float(t))
            concs.append(float(c))
            cfus.append(float(f))
        except (ValueError, TypeError) as exc:
            bad_rows.append(f"row {row_num + 1}: {exc}")

    if bad_rows:
        raise ValueError(f"Non-numeric values in {len(bad_rows)} row(s): {'; '.join(bad_rows[:5])}")
    if not times:
        raise ValueError("No valid data rows found.")

    icts = compute_ict(times, concs)

    sample_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO samples (id, project_id, name, column_map, experiment_id) VALUES (?, ?, ?, ?, ?)",
        (sample_id, project_id, sample_name,
         json.dumps({"time": col_map.time, "concentration": col_map.concentration, "cfu": col_map.cfu}),
         exp_id),
    )
    for i, (t, c, f, ict) in enumerate(zip(times, concs, cfus, icts)):
        conn.execute(
            "INSERT INTO observations (sample_id, time, concentration, cfu, ict, row_index) VALUES (?, ?, ?, ?, ?, ?)",
            (sample_id, t, c, f, ict, i),
        )
    obs_count = len(times)
    conn.execute(
        "INSERT INTO fit_events (id, sample_id, event_type, title, body, metadata) VALUES (?, ?, ?, ?, ?, ?)",
        (str(uuid.uuid4()), sample_id, "sample_created",
         f"Sample imported: {sample_name}",
         f"Sample '{sample_name}' imported via Quick Import with {obs_count} observations.",
         json.dumps({"obs_count": obs_count, "source": "quick_import", "experiment_id": exp_id}))
    )
    return obs_count


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/api/projects/{project_id}/quick-import/preview")
async def quick_import_preview(project_id: str, file: UploadFile = File(...)):
    conn = get_connection()
    proj = conn.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
    conn.close()
    if not proj:
        raise HTTPException(404, "Project not found.")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Failed to open Excel file: {e}")

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        preview = _detect_sheet(ws, name)
        sheets.append(preview)
    wb.close()

    return QuickImportPreview(sheets=sheets)


@router.post("/api/projects/{project_id}/quick-import/execute")
async def quick_import_execute(
    project_id: str,
    file: UploadFile = File(...),
    config: str = Form(...),  # JSON-encoded QuickImportExecuteRequest
):
    try:
        body = QuickImportExecuteRequest.model_validate_json(config)
    except Exception as e:
        raise HTTPException(400, f"Invalid config JSON: {e}")

    if body.project_id != project_id:
        raise HTTPException(400, "project_id mismatch.")

    conn = get_connection()
    proj = conn.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, "Project not found.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        conn.close()
        raise HTTPException(400, f"Failed to open Excel file: {e}")

    experiment_ids: list[int] = []
    total_samples = 0
    total_observations = 0

    try:
        for sheet_cfg in body.sheets:
            if sheet_cfg.sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_cfg.sheet_name}' not found in workbook.")

            ws = wb[sheet_cfg.sheet_name]
            all_rows = [[_cell_str(cell.value) for cell in row] for row in ws.iter_rows()]

            headers = all_rows[sheet_cfg.header_row_index] if all_rows else []

            # Create experiment
            exp_meta = {}
            if sheet_cfg.metadata:
                if sheet_cfg.metadata.experiment_date:
                    exp_meta["experiment_date"] = sheet_cfg.metadata.experiment_date
                if sheet_cfg.metadata.analyst:
                    exp_meta["analyst"] = sheet_cfg.metadata.analyst
                if sheet_cfg.metadata.notes:
                    exp_meta["notes"] = sheet_cfg.metadata.notes
            exp_row = conn.execute(
                "INSERT INTO experiments (project_id, name, metadata) VALUES (?, ?, ?) RETURNING id",
                (project_id, sheet_cfg.experiment_name, json.dumps(exp_meta)),
            ).fetchone()
            exp_id = exp_row[0]
            experiment_ids.append(exp_id)

            if sheet_cfg.data_row_indices is not None:
                # Manual grid: explicit row selection
                data_rows = [all_rows[i] for i in sheet_cfg.data_row_indices
                             if i < len(all_rows) and i != sheet_cfg.header_row_index]
                if sheet_cfg.col_map.group:
                    # Split by group col
                    group_idx = next((i for i, h in enumerate(headers) if h == sheet_cfg.col_map.group), None)
                    if group_idx is not None:
                        groups: dict[str, list] = {}
                        for row in data_rows:
                            gval = row[group_idx].strip() if group_idx < len(row) else ""
                            if not gval:
                                continue
                            groups.setdefault(gval, []).append(row)
                        for gval, grows in groups.items():
                            sname = f"{sheet_cfg.experiment_name} – {gval} mg/L"
                            obs = _insert_sample_quick(conn, project_id, exp_id, sname, sheet_cfg.col_map, grows, headers)
                            total_samples += 1
                            total_observations += obs
                    else:
                        obs = _insert_sample_quick(conn, project_id, exp_id, sheet_cfg.experiment_name, sheet_cfg.col_map, data_rows, headers)
                        total_samples += 1
                        total_observations += obs
                else:
                    obs = _insert_sample_quick(conn, project_id, exp_id, sheet_cfg.experiment_name, sheet_cfg.col_map, data_rows, headers)
                    total_samples += 1
                    total_observations += obs
            else:
                # Auto: skip header row(s), group by group col
                data_rows = all_rows[sheet_cfg.header_row_index + 1:]
                if sheet_cfg.col_map.group:
                    group_idx = next((i for i, h in enumerate(headers) if h == sheet_cfg.col_map.group), None)
                    if group_idx is not None:
                        groups = {}
                        for row in data_rows:
                            gval = row[group_idx].strip() if group_idx < len(row) else ""
                            if not gval or not _is_numeric(gval):
                                continue
                            groups.setdefault(gval, []).append(row)
                        for gval, grows in groups.items():
                            sname = f"{sheet_cfg.experiment_name} – {gval} mg/L"
                            obs = _insert_sample_quick(conn, project_id, exp_id, sname, sheet_cfg.col_map, grows, headers)
                            total_samples += 1
                            total_observations += obs
                    else:
                        obs = _insert_sample_quick(conn, project_id, exp_id, sheet_cfg.experiment_name, sheet_cfg.col_map, data_rows, headers)
                        total_samples += 1
                        total_observations += obs
                else:
                    obs = _insert_sample_quick(conn, project_id, exp_id, sheet_cfg.experiment_name, sheet_cfg.col_map, data_rows, headers)
                    total_samples += 1
                    total_observations += obs

        conn.commit()
        wb.close()
    except Exception as e:
        conn.rollback()
        wb.close()
        conn.close()
        raise HTTPException(500, f"Import failed: {e}")

    conn.close()
    return QuickImportResult(
        experiment_ids=experiment_ids,
        first_experiment_id=experiment_ids[0] if experiment_ids else 0,
        total_samples=total_samples,
        total_observations=total_observations,
    )
