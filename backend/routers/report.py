import json
import io
import math
import os
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from ..db.database import get_connection

router = APIRouter(prefix="/api", tags=["report"])

# Matplotlib must use non-interactive backend before any other import
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np


class CalculatedNRow(BaseModel):
    ict: float
    values: List[float]


class ReportRequest(BaseModel):
    fit_id: str
    chart_png_base64: Optional[str] = None
    sample_names: List[str] = []
    sample_ids: Optional[List[str]] = None
    calculated_n: List[CalculatedNRow] = []
    experiment_id: Optional[int] = None


SAMPLE_COLORS = ["#3b82f6", "#10b981", "#f59e0b", "#ef4444",
                 "#8b5cf6", "#06b6d4", "#ec4899", "#84cc16"]


def style_header(cell, bg_color="1e3a5f"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def section_title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color="1e3a5f")
    c.fill = PatternFill("solid", fgColor="dbeafe")
    ws.row_dimensions[row].height = 20
    return row + 1


def auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


def fmt_param(v: float) -> str:
    if v == 0:
        return "0"
    a = abs(v)
    if a < 0.001 or a >= 100000:
        return f"{v:.4e}"
    if a < 0.1:
        return f"{v:.6f}"
    return f"{v:.4f}"


def build_equation(model_id: str, params: list) -> str:
    p = {item.get("name", ""): item.get("value", 0) for item in params}
    if model_id == "two_population_ict":
        beta = fmt_param(p.get("beta", 0))
        kd = fmt_param(p.get("kd", 0))
        kp = fmt_param(p.get("kp", 0))
        m = fmt_param(p.get("m", 1))
        return f"N/N0 = (1 - {beta}) * exp(-{kd} * ICT^{m}) + {beta} * exp(-{kp} * ICT)"
    # Generic fallback
    return "  |  ".join(f"{item.get('name')} = {fmt_param(item.get('value', 0))}" for item in params)


def predict_frac(ict, params):
    try:
        beta = params[0]["value"]
        kd = params[1]["value"]
        kp = params[2]["value"]
        m = params[3]["value"]
        return (1 - beta) * math.exp(-kd * max(ict, 0) ** m) + beta * math.exp(-kp * ict)
    except Exception:
        return 1.0


def build_fit_plot(obs_by_sample, all_sample_names, all_sample_ids, params) -> bytes:
    """Generate the fit model plot as PNG bytes with white background."""
    fig, ax = plt.subplots(figsize=(9, 5.5), dpi=120)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    all_ict = []
    for sid, sname in zip(all_sample_ids, all_sample_names):
        obs = obs_by_sample.get(sid, [])
        if not obs:
            continue
        color = SAMPLE_COLORS[all_sample_ids.index(sid) % len(SAMPLE_COLORS)]
        xs = [r["ict"] or 0 for r in obs]
        ys = [r["cfu"] for r in obs]
        ax.scatter(xs, ys, color=color, label=sname, s=50, zorder=5)
        all_ict.extend(xs)

    # Prediction curve using median N0 across samples
    n0_values = []
    for sid in all_sample_ids:
        obs = obs_by_sample.get(sid, [])
        if obs:
            n0_values.append(max(r["cfu"] for r in obs))
    if not n0_values:
        n0_values = [1]
    n0 = sorted(n0_values)[len(n0_values) // 2]

    x_curve = np.linspace(0, 60, 300)
    y_curve = [predict_frac(x, params) * n0 for x in x_curve]
    ax.plot(x_curve, y_curve, color="#f59e0b", linewidth=2.5, label="Model fit", zorder=4)

    # Detection limit
    ax.axhline(1, color="#6b7280", linestyle="--", linewidth=1, alpha=0.6, label="Detection limit (1 CFU)")

    ax.set_yscale("log")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v, _: f"{v:,.0f}" if v >= 1 else f"{v:.1e}"
    ))
    ax.set_xlim(0, 60)
    ax.set_xlabel("Total ICT (mg·min/L)", fontsize=11)
    ax.set_ylabel("Viable Microbes (CFU/100mL)", fontsize=11)
    ax.set_title("Disinfection Model Fit", fontsize=13, fontweight="bold")
    ax.legend(framealpha=0.9, fontsize=9)
    ax.grid(True, which="major", color="#e5e7eb", linewidth=0.7)
    ax.grid(True, which="minor", color="#f3f4f6", linewidth=0.4)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def build_diagnostic_plots(diag: dict) -> bytes:
    """Generate 2×2 diagnostic chart panel as PNG bytes."""
    fig, axes = plt.subplots(2, 2, figsize=(12, 8), dpi=110)
    fig.patch.set_facecolor("white")
    fig.suptitle("Model Diagnostics", fontsize=14, fontweight="bold", y=1.01)

    ict = diag.get("ict_values", [])
    std_res = diag.get("standardized_residuals", [])
    qq_t = diag.get("qq_theoretical", [])
    qq_s = diag.get("qq_sample", [])
    cooks = diag.get("cooks_distance", [])

    # 1 — Residuals vs ICT
    ax = axes[0, 0]
    ax.set_facecolor("white")
    ax.scatter(ict, std_res, color="#3b82f6", s=50, zorder=5)
    ax.axhline(0, color="#ef4444", linestyle="--", linewidth=1.2)
    ax.set_xlabel("ICT (mg·min/L)", fontsize=10)
    ax.set_ylabel("Standardized Residual", fontsize=10)
    ax.set_title("Residuals vs ICT", fontsize=11, fontweight="bold")
    ax.grid(True, color="#e5e7eb", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # 2 — Q-Q Plot
    ax = axes[0, 1]
    ax.set_facecolor("white")
    ax.scatter(qq_t, qq_s, color="#8b5cf6", s=50, zorder=5)
    if qq_t:
        lim = [min(min(qq_t), min(qq_s)), max(max(qq_t), max(qq_s))]
        ax.plot(lim, lim, color="#ef4444", linestyle="--", linewidth=1.2)
    ax.set_xlabel("Theoretical Quantiles", fontsize=10)
    ax.set_ylabel("Sample Quantiles", fontsize=10)
    ax.set_title("Q-Q Plot", fontsize=11, fontweight="bold")
    ax.grid(True, color="#e5e7eb", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # 3 — Histogram of residuals
    ax = axes[1, 0]
    ax.set_facecolor("white")
    if std_res:
        ax.hist(std_res, bins=max(5, len(std_res) // 2), color="#10b981",
                alpha=0.8, edgecolor="white")
    ax.set_xlabel("Standardized Residual", fontsize=10)
    ax.set_ylabel("Count", fontsize=10)
    ax.set_title("Residual Histogram", fontsize=11, fontweight="bold")
    ax.grid(True, axis="y", color="#e5e7eb", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # 4 — Cook's Distance
    ax = axes[1, 1]
    ax.set_facecolor("white")
    if cooks:
        colors = ["#ef4444" if v > 1 else "#f59e0b" if v > 0.5 else "#3b82f6"
                  for v in cooks]
        ax.bar(range(1, len(cooks) + 1), cooks, color=colors)
        ax.axhline(1, color="#ef4444", linestyle="--", linewidth=1.2, label="Threshold (1.0)")
        ax.legend(fontsize=9)
    ax.set_xlabel("Observation Index", fontsize=10)
    ax.set_ylabel("Cook's Distance", fontsize=10)
    ax.set_title("Cook's Distance", fontsize=11, fontweight="bold")
    ax.grid(True, axis="y", color="#e5e7eb", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


@router.post("/samples/{sample_id}/report")
def generate_report(sample_id: str, body: ReportRequest):
    conn = get_connection()

    sample = conn.execute("SELECT * FROM samples WHERE id = ?", (sample_id,)).fetchone()
    if not sample:
        conn.close()
        raise HTTPException(404, "Sample not found")
    sample = dict(sample)

    project = conn.execute("SELECT * FROM projects WHERE id = ?", (sample["project_id"],)).fetchone()
    project = dict(project) if project else {"name": "Unknown", "description": ""}

    fit = conn.execute("SELECT * FROM model_fits WHERE id = ?", (body.fit_id,)).fetchone()
    if not fit:
        conn.close()
        raise HTTPException(404, "Fit not found")
    fit = dict(fit)
    params = json.loads(fit["parameters"])
    stats = json.loads(fit["statistics"])
    qs = json.loads(fit["quality_score"])
    diag = json.loads(fit["diagnostics"])

    all_sample_ids = body.sample_ids if body.sample_ids else [sample_id]
    all_sample_names = body.sample_names if body.sample_names else [sample["name"]]

    obs_by_sample: dict[str, list] = {}
    for sid in all_sample_ids:
        rows = conn.execute(
            "SELECT time, concentration, cfu, ict FROM observations WHERE sample_id = ? ORDER BY time",
            (sid,),
        ).fetchall()
        obs_by_sample[sid] = [dict(r) for r in rows]

    placeholders = ",".join("?" * len(all_sample_ids))
    events = conn.execute(
        f"SELECT * FROM fit_events WHERE sample_id IN ({placeholders}) ORDER BY created_at ASC",
        all_sample_ids,
    ).fetchall()
    events = [dict(e) for e in events]
    sid_to_name = dict(zip(all_sample_ids, all_sample_names))
    for ev in events:
        ev["sample_name"] = sid_to_name.get(ev.get("sample_id", ""), ev.get("sample_id", ""))
        ev["metadata_parsed"] = json.loads(ev["metadata"]) if ev.get("metadata") else {}

    # Fetch experiment info before closing conn
    experiment_name = None
    output_path = None
    if body.experiment_id:
        exp_row = conn.execute(
            "SELECT e.name, p.output_path FROM experiments e "
            "JOIN projects p ON p.id = e.project_id "
            "WHERE e.id = ?",
            (body.experiment_id,),
        ).fetchone()
        if exp_row:
            experiment_name = exp_row["name"]
            output_path = exp_row["output_path"]

    conn.close()

    # Build obs vs predicted
    obs_pred_rows = []
    for sid, sname in zip(all_sample_ids, all_sample_names):
        sample_obs = obs_by_sample.get(sid, [])
        if not sample_obs:
            continue
        n0 = max(r["cfu"] for r in sample_obs)
        for r in sample_obs:
            ict = r["ict"] or 0
            frac = predict_frac(ict, params)
            pred_n = frac * n0
            residual = r["cfu"] - pred_n
            pct_err = (residual / r["cfu"] * 100) if r["cfu"] != 0 else 0
            obs_pred_rows.append({
                "sample": sname,
                "time": r["time"],
                "concentration": r["concentration"],
                "ict": ict,
                "observed": r["cfu"],
                "predicted": pred_n,
                "residual": residual,
                "pct_error": pct_err,
            })

    # Generate plots server-side
    fit_plot_png = build_fit_plot(obs_by_sample, all_sample_names, all_sample_ids, params)
    diag_plot_png = build_diagnostic_plots(diag)

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    report_date = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    # ── ① Cover ─────────────────────────────────────────────────────────────
    ws_cover = wb.create_sheet("① Cover")
    ws_cover.column_dimensions["A"].width = 30
    ws_cover.column_dimensions["B"].width = 55
    cover_data = [
        ("Report Title", "Disinfection Benchmark Modeling Report"),
        ("Project", project["name"]),
        ("Description", project.get("description") or ""),
        ("Samples Included", ", ".join(all_sample_names)),
        ("Report Generated", report_date),
        ("Model", fit["model_id"].replace("_", " ").title()),
        ("Fit ID", body.fit_id),
        ("R² (log)", f"{stats.get('r_squared', 0):.4f}" if isinstance(stats.get("r_squared"), float) else "N/A"),
        ("Quality Score", f"{qs.get('score', 0):.0f} / 100 ({qs.get('rating', '')})" if isinstance(qs.get("score"), (int, float)) else "N/A"),
        ("Total Observations", len(obs_pred_rows)),
        ("Converged", "Yes" if stats.get("converged") else "No"),
    ]
    for i, (label, value) in enumerate(cover_data, start=2):
        ws_cover.cell(row=i, column=1, value=label).font = Font(bold=True, color="1e3a5f")
        ws_cover.cell(row=i, column=2, value=str(value))
        ws_cover.row_dimensions[i].height = 18

    # ── ② Plot ──────────────────────────────────────────────────────────────
    ws_plot = wb.create_sheet("② Plot")
    ws_plot.cell(row=1, column=1, value="Project").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=1, column=2, value=project["name"])
    ws_plot.cell(row=2, column=1, value="Sample(s)").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=2, column=2, value=", ".join(all_sample_names))
    ws_plot.cell(row=3, column=1, value="Model").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=3, column=2, value=fit["model_id"].replace("_", " ").title())
    ws_plot.cell(row=4, column=1, value="R² (log)").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=4, column=2, value=round(stats.get("r_squared", 0), 4))
    ws_plot.cell(row=5, column=1, value="Quality Score").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=5, column=2, value=f"{qs.get('score', 0):.0f} / 100 ({qs.get('rating', '')})")
    ws_plot.cell(row=6, column=1, value="Generated").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=6, column=2, value=report_date)
    # Parameters — one row each starting at row 7
    plot_row = 7
    for p in params:
        ws_plot.cell(row=plot_row, column=1, value=p.get("name", "")).font = Font(bold=True, color="1e3a5f")
        val_str = fmt_param(p.get("value", 0))
        if p.get("std_error"):
            val_str += f"  ±{fmt_param(p['std_error'])}"
        ws_plot.cell(row=plot_row, column=2, value=val_str)
        plot_row += 1
    # Blank row
    plot_row += 1
    # Fitted equation
    ws_plot.cell(row=plot_row, column=1, value="Fitted Equation").font = Font(bold=True, color="1e3a5f")
    ws_plot.cell(row=plot_row, column=2, value=build_equation(fit["model_id"], params))
    ws_plot.cell(row=plot_row, column=2).font = Font(name="Courier New", size=10, color="1e3a5f")
    # 2 blank rows then image
    plot_row += 3
    ws_plot.column_dimensions["A"].width = 18
    ws_plot.column_dimensions["B"].width = 90
    img = openpyxl.drawing.image.Image(io.BytesIO(fit_plot_png))
    img.anchor = f"A{plot_row}"
    img.width = 720
    img.height = 450
    ws_plot.add_image(img)

    # ── ③ Obs vs Predicted ──────────────────────────────────────────────────
    ws_ov = wb.create_sheet("③ Obs vs Predicted")
    ws_ov.freeze_panes = "A2"
    headers_ov = ["Sample", "Time (min)", "Concentration (mg/L)", "ICT (mg·min/L)",
                  "Observed N (CFU/100mL)", "Predicted N (CFU/100mL)", "Residual", "% Error"]
    for ci, h in enumerate(headers_ov, 1):
        style_header(ws_ov.cell(row=1, column=ci, value=h))
    for ri, r in enumerate(obs_pred_rows, 2):
        ws_ov.cell(row=ri, column=1, value=r["sample"])
        ws_ov.cell(row=ri, column=2, value=round(r["time"], 4))
        ws_ov.cell(row=ri, column=3, value=round(r["concentration"], 4))
        ws_ov.cell(row=ri, column=3).number_format = "0.0000"
        ws_ov.cell(row=ri, column=4, value=round(r["ict"], 4))
        ws_ov.cell(row=ri, column=4).number_format = "0.0000"
        ws_ov.cell(row=ri, column=5, value=round(r["observed"], 2))
        ws_ov.cell(row=ri, column=5).number_format = "#,##0.00"
        ws_ov.cell(row=ri, column=6, value=round(r["predicted"], 2))
        ws_ov.cell(row=ri, column=6).number_format = "#,##0.00"
        ws_ov.cell(row=ri, column=7, value=round(r["residual"], 2))
        ws_ov.cell(row=ri, column=7).number_format = "#,##0.00"
        ws_ov.cell(row=ri, column=8, value=round(r["pct_error"], 2))
        ws_ov.cell(row=ri, column=8).number_format = "0.00"
    if obs_pred_rows:
        mean_abs_pct = sum(abs(r["pct_error"]) for r in obs_pred_rows) / len(obs_pred_rows)
        rmse_val = (sum(r["residual"] ** 2 for r in obs_pred_rows) / len(obs_pred_rows)) ** 0.5
        sr = len(obs_pred_rows) + 3
        ws_ov.cell(row=sr, column=5, value="Mean Abs % Error:").font = Font(bold=True)
        ws_ov.cell(row=sr, column=8, value=round(mean_abs_pct, 2)).number_format = "0.00"
        ws_ov.cell(row=sr + 1, column=5, value="RMSE (CFU/100mL):").font = Font(bold=True)
        ws_ov.cell(row=sr + 1, column=8, value=round(rmse_val, 2)).number_format = "#,##0.00"
    auto_width(ws_ov)

    # ── ④ Calculated N ──────────────────────────────────────────────────────
    ws_cn = wb.create_sheet("④ Calculated N")
    # Parameter summary at top
    ws_cn.cell(row=1, column=1, value="Model").font = Font(bold=True, color="1e3a5f")
    ws_cn.cell(row=1, column=2, value=fit["model_id"].replace("_", " ").title())
    ws_cn.cell(row=2, column=1, value="R² (log)").font = Font(bold=True, color="1e3a5f")
    ws_cn.cell(row=2, column=2, value=round(stats.get("r_squared", 0), 4))
    for pi, p in enumerate(params, 3):
        ws_cn.cell(row=pi, column=1, value=p.get("name", "")).font = Font(bold=True, color="1e3a5f")
        val_str = f"{p.get('value', 0):.6g}"
        if p.get("std_error"):
            val_str += f"  ±{p.get('std_error'):.2g}"
        ws_cn.cell(row=pi, column=2, value=val_str)
    data_start_row = len(params) + 4  # blank row gap then header
    ws_cn.freeze_panes = f"A{data_start_row + 1}"
    cn_headers = ["ICT (mg·min/L)"] + [f"{n} (CFU/100mL)" for n in (all_sample_names or ["N"])]
    for ci, h in enumerate(cn_headers, 1):
        style_header(ws_cn.cell(row=data_start_row, column=ci, value=h))
    for ri, cn_row in enumerate(body.calculated_n, data_start_row + 1):
        ws_cn.cell(row=ri, column=1, value=cn_row.ict)
        ws_cn.cell(row=ri, column=1).number_format = "0.0"
        for ci, val in enumerate(cn_row.values, 2):
            ws_cn.cell(row=ri, column=ci, value=round(val, 2))
            ws_cn.cell(row=ri, column=ci).number_format = "#,##0.00"
    auto_width(ws_cn)

    # ── ⑤ Parameters ────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("⑤ Parameters")
    ws_p.freeze_panes = "A2"
    for ci, h in enumerate(["Parameter", "Value", "Std Error", "CI Lower", "CI Upper"], 1):
        style_header(ws_p.cell(row=1, column=ci, value=h))
    for ri, p in enumerate(params, 2):
        ws_p.cell(row=ri, column=1, value=p.get("name", ""))
        ws_p.cell(row=ri, column=2, value=round(p.get("value", 0), 8))
        ws_p.cell(row=ri, column=3, value=round(p.get("std_error", 0) or 0, 8))
        ws_p.cell(row=ri, column=4, value=round(p.get("ci_lower", 0) or 0, 8) if p.get("ci_lower") is not None else "N/A")
        ws_p.cell(row=ri, column=5, value=round(p.get("ci_upper", 0) or 0, 8) if p.get("ci_upper") is not None else "N/A")
    auto_width(ws_p)

    # ── ⑥ Diagnostics ───────────────────────────────────────────────────────
    ws_diag = wb.create_sheet("⑥ Diagnostics")
    ws_diag.cell(row=1, column=1, value="Model Diagnostics — 4-panel chart").font = Font(bold=True, size=13, color="1e3a5f")
    ws_diag.cell(row=2, column=1, value="Charts: Residuals vs ICT · Q-Q Plot · Residual Histogram · Cook's Distance").font = Font(italic=True, color="475569")
    diag_img = openpyxl.drawing.image.Image(io.BytesIO(diag_plot_png))
    diag_img.anchor = "A4"
    diag_img.width = 900
    diag_img.height = 600
    ws_diag.add_image(diag_img)

    # ── ⑦ Statistics & Quality ──────────────────────────────────────────────
    ws_sq = wb.create_sheet("⑦ Statistics & Quality")
    ws_sq.freeze_panes = "A2"
    ws_sq.column_dimensions["A"].width = 32
    ws_sq.column_dimensions["B"].width = 25

    # Statistics section
    row = section_title(ws_sq, 1, "Fit Statistics")
    style_header(ws_sq.cell(row=row, column=1, value="Metric"))
    style_header(ws_sq.cell(row=row, column=2, value="Value"))
    row += 1
    stat_rows = [
        ("R² (log space)", stats.get("r_squared")),
        ("Adj. R² (log space)", stats.get("adj_r_squared")),
        ("RMSE (log₁₀)", stats.get("rmse")),
        ("MAE (log₁₀)", stats.get("mae")),
        ("SSE", stats.get("sse")),
        ("MSE", stats.get("mse")),
        ("AIC", stats.get("aic")),
        ("BIC", stats.get("bic")),
        ("Log-Likelihood", stats.get("log_likelihood")),
        ("Observations", stats.get("n_obs")),
        ("Parameters", stats.get("n_params")),
        ("Degrees of Freedom", (stats.get("n_obs") or 0) - (stats.get("n_params") or 0)),
        ("Converged", "Yes" if stats.get("converged") else "No"),
        ("Iterations", stats.get("n_iterations")),
    ]
    for label, val in stat_rows:
        ws_sq.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_sq.cell(row=row, column=2, value=val if val is not None else "N/A")
        row += 1

    row += 1  # blank row
    row = section_title(ws_sq, row, "Quality Score")
    style_header(ws_sq.cell(row=row, column=1, value="Category"))
    style_header(ws_sq.cell(row=row, column=2, value="Value"))
    row += 1
    q_rows = [
        ("Overall Score", f"{qs.get('score', 0):.1f} / 100"),
        ("Rating", qs.get("rating", "")),
        ("Data Quality", qs.get("data_quality", "")),
        ("Optimization Quality", qs.get("optimization_quality", "")),
        ("Statistical Quality", qs.get("statistical_quality", "")),
        ("Scientific Quality", qs.get("scientific_quality", "")),
    ]
    for label, val in q_rows:
        ws_sq.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_sq.cell(row=row, column=2, value=str(val))
        row += 1

    row += 1
    ws_sq.cell(row=row, column=1, value="Strengths").font = Font(bold=True, color="1e8c1e")
    row += 1
    for s in qs.get("strengths", []):
        ws_sq.cell(row=row, column=1, value=f"✓ {s}")
        row += 1

    row += 1
    ws_sq.cell(row=row, column=1, value="Weaknesses").font = Font(bold=True, color="c0392b")
    row += 1
    for w in qs.get("weaknesses", []):
        ws_sq.cell(row=row, column=1, value=f"✗ {w}")
        row += 1

    row += 1
    ws_sq.cell(row=row, column=1, value="Deductions").font = Font(bold=True)
    row += 1
    for d in qs.get("deductions", []):
        ws_sq.cell(row=row, column=1, value=d.get("category", ""))
        ws_sq.cell(row=row, column=2, value=d.get("reason", ""))
        ws_sq.cell(row=row, column=3, value=f"-{d.get('points', 0)} pts")
        row += 1

    # ── ⑧ Process Journal ───────────────────────────────────────────────────
    ws_j = wb.create_sheet("⑧ Process Journal")
    ws_j.freeze_panes = "A2"
    headers_j = ["Timestamp", "Sample", "Event Type", "Title", "Description", "R²", "Quality Score", "Converged"]
    for ci, h in enumerate(headers_j, 1):
        style_header(ws_j.cell(row=1, column=ci, value=h))
    for ri, ev in enumerate(events, 2):
        md = ev.get("metadata_parsed", {})
        ws_j.cell(row=ri, column=1, value=ev.get("created_at", ""))
        ws_j.cell(row=ri, column=2, value=ev.get("sample_name", ""))
        ws_j.cell(row=ri, column=3, value=ev.get("event_type", ""))
        ws_j.cell(row=ri, column=4, value=ev.get("title", ""))
        ws_j.cell(row=ri, column=5, value=ev.get("body", ""))
        ws_j.cell(row=ri, column=6, value=round(md["r_squared"], 4) if "r_squared" in md else "")
        ws_j.cell(row=ri, column=7, value=round(md["score"], 1) if "score" in md else "")
        ws_j.cell(row=ri, column=8, value="Yes" if md.get("converged") else ("No" if "converged" in md else ""))
        ws_j.row_dimensions[ri].height = 40
        ws_j.cell(row=ri, column=5).alignment = Alignment(wrap_text=True)
    auto_width(ws_j)

    # ── Determine filename ───────────────────────────────────────────────────
    timestamp = datetime.utcnow().strftime("%m%d%Y-%H%M")

    if experiment_name:
        safe_stem = experiment_name.replace(" ", "_").replace("/", "-")
    else:
        safe_stem = sample["name"].replace(" ", "_").replace("/", "-")

    filename = f"{safe_stem}-{timestamp}.xlsx"

    # ── Build workbook bytes ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    # ── Optionally save to project output_path ───────────────────────────────
    if output_path:
        try:
            dest = os.path.join(output_path, filename)
            with open(dest, "wb") as fh:
                fh.write(xlsx_bytes)
        except Exception:
            pass  # disk save is best-effort; download always proceeds

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
